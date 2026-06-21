import json
import os
import re
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

EU_PAGE = "https://energy.ec.europa.eu/data-and-analysis/weekly-oil-bulletin_en"
OUT = Path("fuel-prices.json")
ROUTE_COUNTRIES = {
    "Germany": {"code": "DE", "country": "Deutschland"},
    "Austria": {"code": "AT", "country": "Österreich"},
    "Hungary": {"code": "HU", "country": "Ungarn"},
    "Romania": {"code": "RO", "country": "Rumänien"},
    "Bulgaria": {"code": "BG", "country": "Bulgarien"},
}


def round_price(value):
    if value is None:
        return None
    try:
        value = float(value)
    except (TypeError, ValueError):
        return None
    if value <= 0:
        return None
    # EU Oil Bulletin is often EUR/1000 litres. Convert to EUR/litre.
    if value > 20:
        value = value / 1000
    return round(value, 3)


def get_latest_xlsx_url():
    response = requests.get(EU_PAGE, timeout=30)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, "html.parser")

    best = None
    for link in soup.find_all("a", href=True):
        href = link["href"]
        text = link.get_text(" ", strip=True).lower()
        parent_text = link.find_parent().get_text(" ", strip=True).lower() if link.find_parent() else text
        href_lower = href.lower()
        block = f"{text} {parent_text} {href_lower}"
        if "document/download" in href_lower and "xlsx" in block and "tax" in block:
            best = urljoin(EU_PAGE, href)
            if "with" in block and "without" not in block:
                return best
    if best:
        return best
    raise RuntimeError("Kein XLSX-Link für Preise mit Steuern gefunden")


def find_columns(rows, country_row_index):
    start = max(0, country_row_index - 10)
    header_rows = rows[start:country_row_index]
    max_cols = max(len(row) for row in rows if row)
    headers = []
    for col in range(max_cols):
        parts = []
        for row in header_rows:
            if col < len(row) and row[col] is not None:
                parts.append(str(row[col]).strip())
        headers.append(" ".join(parts).lower())

    gasoline_col = None
    diesel_col = None
    for index, header in enumerate(headers):
        compact = re.sub(r"\s+", " ", header)
        if gasoline_col is None and ("euro-super 95" in compact or "eurosuper 95" in compact or ("super" in compact and "95" in compact)):
            gasoline_col = index
        if diesel_col is None and ("automotive gas oil" in compact or "road diesel" in compact or "diesel" in compact):
            diesel_col = index
    return gasoline_col, diesel_col


def parse_eu_xlsx(path):
    wb = load_workbook(path, data_only=True)
    results = {}

    for ws in wb.worksheets:
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        for r_index, row in enumerate(rows):
            row_text = " ".join(str(cell).strip() for cell in row if cell is not None)
            for english_name, meta in ROUTE_COUNTRIES.items():
                if english_name not in row_text:
                    continue
                gasoline_col, diesel_col = find_columns(rows, r_index)

                numeric_values = [cell for cell in row if isinstance(cell, (int, float))]
                gasoline = round_price(row[gasoline_col]) if gasoline_col is not None and gasoline_col < len(row) else None
                diesel = round_price(row[diesel_col]) if diesel_col is not None and diesel_col < len(row) else None

                # Fallback: in the EU bulletin, gasoline and diesel are usually among the first fuel price numbers.
                if gasoline is None and numeric_values:
                    gasoline = round_price(numeric_values[0])
                if diesel is None and len(numeric_values) > 1:
                    diesel = round_price(numeric_values[1])

                if gasoline is not None or diesel is not None:
                    results[meta["code"]] = {
                        "code": meta["code"],
                        "country": meta["country"],
                        "diesel_eur": diesel,
                        "gasoline95_eur": gasoline,
                        "source": "European Commission Weekly Oil Bulletin",
                    }
    return results


def optional_turkey(existing_country):
    """Optional hook for a paid/free Turkey endpoint.

    Set a repository secret TURKEY_FUEL_API_URL that returns JSON. The script accepts common keys:
    diesel_eur, gasoline95_eur, diesel, gasoline95, benzin, motorin.
    """
    base = {
        "code": "TR",
        "country": "Türkei",
        "diesel_eur": None,
        "gasoline95_eur": None,
        "source": "Keine freie offizielle EU-Quelle; optional über TURKEY_FUEL_API_URL im Workflow",
    }
    if isinstance(existing_country, dict):
        base.update(existing_country)

    api_url = os.environ.get("TURKEY_FUEL_API_URL")
    if not api_url:
        return base

    try:
        response = requests.get(api_url, timeout=30)
        response.raise_for_status()
        data = response.json()
        if isinstance(data, list) and data:
            data = data[0]
        if not isinstance(data, dict):
            return base

        diesel = data.get("diesel_eur") or data.get("diesel") or data.get("motorin")
        gasoline = data.get("gasoline95_eur") or data.get("gasoline95") or data.get("benzin")
        base["diesel_eur"] = round_price(diesel)
        base["gasoline95_eur"] = round_price(gasoline)
        base["source"] = "TURKEY_FUEL_API_URL"
    except Exception as exc:
        base["source"] = f"Turkey API nicht geladen: {exc}"
    return base


def main():
    previous = {}
    if OUT.exists():
        try:
            previous = json.loads(OUT.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            previous = {}

    previous_by_code = {item.get("code"): item for item in previous.get("countries", []) if isinstance(item, dict)}

    url = get_latest_xlsx_url()
    xlsx = Path("weekly-oil-bulletin.xlsx")
    response = requests.get(url, timeout=60)
    response.raise_for_status()
    xlsx.write_bytes(response.content)

    eu_prices = parse_eu_xlsx(xlsx)

    countries = []
    for code, country in [
        ("DE", "Deutschland"),
        ("AT", "Österreich"),
        ("HU", "Ungarn"),
        ("RO", "Rumänien"),
        ("BG", "Bulgarien"),
    ]:
        item = eu_prices.get(code) or previous_by_code.get(code) or {
            "code": code,
            "country": country,
            "diesel_eur": None,
            "gasoline95_eur": None,
            "source": "European Commission Weekly Oil Bulletin",
        }
        countries.append(item)

    countries.append(optional_turkey(previous_by_code.get("TR")))

    output = {
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "source": "EU: European Commission Weekly Oil Bulletin. Turkey: optional TURKEY_FUEL_API_URL.",
        "source_url": EU_PAGE,
        "status": "updated",
        "unit": "EUR/L",
        "note": "EU-Werte sind nationale Durchschnittspreise mit Steuern, nicht einzelne Tankstellenpreise.",
        "countries": countries,
    }
    OUT.write_text(json.dumps(output, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
